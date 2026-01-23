from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

SEARCH_TEXT = "coffee shops near GST Road Chrompet"


def extract_coffee_shops():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        print("🌐 Loading Google Maps...")
        page.goto("https://www.google.com/maps")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(5000)

        print("🔍 Searching for coffee shops...")
        search_box = page.locator('input[name="q"]').first
        search_box.click()
        page.wait_for_timeout(1000)
        search_box.fill(SEARCH_TEXT)
        page.wait_for_timeout(1000)
        page.keyboard.press("Enter")

        print("⏳ Waiting for results...")
        page.wait_for_timeout(7000)

        # Scroll to load more results
        try:
            feed = page.locator('div[role="feed"]')
            print("📜 Scrolling to load more results...")
            for scroll in range(5):
                feed.evaluate("el => el.scrollTo(0, el.scrollHeight)")
                page.wait_for_timeout(2000)
                print(f"   Scroll {scroll + 1}/5")
        except:
            print("⚠️ Could not scroll feed")

        # Get all results
        results = page.locator('div[role="article"]')
        count = results.count()

        print(f"\n✅ Found {count} results\n")
        print("="*100)
        print(f"🗺️  RESULTS FOR: {SEARCH_TEXT}")
        print("="*100 + "\n")

        excel_data = []

        for i in range(count):
            article = results.nth(i)

            # Extract name
            try:
                name = article.locator(
                    'div.fontHeadlineSmall').inner_text(timeout=2000)
            except:
                name = "N/A"

            # Extract rating
            try:
                rating = article.locator(
                    'span.MW4etd').inner_text(timeout=2000)
            except:
                rating = "N/A"

            # Extract review count
            try:
                reviews_text = article.locator(
                    'span.UY7F9').first.inner_text(timeout=2000)
                reviews = reviews_text.strip('()')
            except:
                reviews = "N/A"

            # Extract price range
            try:
                price_elements = article.locator('span.UY7F9')
                if price_elements.count() > 1:
                    price = price_elements.nth(1).inner_text(timeout=2000)
                else:
                    price = "N/A"
            except:
                price = "N/A"

            # Extract full text and parse
            try:
                full_text = article.inner_text(timeout=2000)
                lines = [line.strip()
                         for line in full_text.split('\n') if line.strip()]

                category = "N/A"
                address = "N/A"
                hours = "N/A"
                services = []

                for line in lines:
                    # Extract category and address
                    if '·' in line and ('shop' in line.lower() or 'cafe' in line.lower() or 'coffee' in line.lower()):
                        parts = [p.strip() for p in line.split('·')]
                        if len(parts) >= 1:
                            category = parts[0]
                        if len(parts) >= 3:
                            for part in parts[1:]:
                                if any(keyword in part.lower() for keyword in ['road', 'rd', 'street', 'nagar', 'floor', 'no', 'complex']):
                                    address = part
                                    break

                    # Extract hours
                    if any(keyword in line for keyword in ['Open', 'Closed', 'Closes', '24 hours']):
                        hours = line

                    # Extract services
                    if line in ['Dine-in', 'Takeaway', 'Delivery', 'Drive-through', 'No delivery']:
                        services.append(line)

                services_str = ', '.join(services) if services else "N/A"

            except:
                category = "N/A"
                address = "N/A"
                hours = "N/A"
                services_str = "N/A"

            # Get URL
            try:
                link = article.locator('a[href*="/maps/place"]').first
                url = link.get_attribute('href')
            except:
                url = "N/A"

            # Store data
            row = {
                'Name': name,
                'Rating': rating,
                'Reviews': reviews,
                'Price': price,
                'Category': category,
                'Address': address,
                'Hours': hours,
                'Services': services_str,
                'URL': url
            }
            excel_data.append(row)

            # Print formatted output
            print(f"{i+1}. 📍 {name}")
            print(f"   ⭐ {rating} ({reviews} reviews) | 💰 {price}")
            print(f"   🏪 {category}")
            print(f"   📌 {address}")
            print(f"   🕒 {hours}")
            if services_str != "N/A":
                print(f"   🍽️  {services_str}")
            print(f"   🔗 {url[:70]}...")
            print()

        # Create Excel file
        excel_filename = "coffee_shops_chrompet.xlsx"
        create_excel(excel_data, excel_filename)

        print("="*100)
        print(f"✅ Successfully extracted {len(excel_data)} coffee shops!")
        print(f"📊 Data saved to: {excel_filename}")
        print("="*100)

        # Summary statistics
        rated_shops = [r for r in excel_data if r['Rating'] != 'N/A']
        if rated_shops:
            avg_rating = sum(float(r['Rating'])
                             for r in rated_shops) / len(rated_shops)
            print(f"\n📈 Average Rating: {avg_rating:.2f} ⭐")
            print(
                f"🏆 Highest Rated: {max(rated_shops, key=lambda x: float(x['Rating']))['Name']} ({max(float(r['Rating']) for r in rated_shops)}⭐)")

        input("\nPress Enter to close browser...")
        browser.close()


def create_excel(data, filename):
    """Create a formatted Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Coffee Shops"

    # Define headers
    headers = ['S.No', 'Name', 'Rating', 'Reviews', 'Price',
               'Category', 'Address', 'Hours', 'Services', 'URL']

    # Header styling
    header_fill = PatternFill(start_color="4472C4",
                              end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    for row_num, row_data in enumerate(data, 2):
        # Serial number
        ws.cell(row=row_num, column=1, value=row_num - 1)

        # Other columns
        ws.cell(row=row_num, column=2, value=row_data['Name'])
        ws.cell(row=row_num, column=3, value=row_data['Rating'])
        ws.cell(row=row_num, column=4, value=row_data['Reviews'])
        ws.cell(row=row_num, column=5, value=row_data['Price'])
        ws.cell(row=row_num, column=6, value=row_data['Category'])
        ws.cell(row=row_num, column=7, value=row_data['Address'])
        ws.cell(row=row_num, column=8, value=row_data['Hours'])
        ws.cell(row=row_num, column=9, value=row_data['Services'])

        # URL as hyperlink
        url_cell = ws.cell(row=row_num, column=10, value=row_data['URL'])
        if row_data['URL'] != 'N/A':
            url_cell.hyperlink = row_data['URL']
            url_cell.font = Font(color="0000FF", underline="single")

        # Apply borders to all cells
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border

    # Adjust column widths
    column_widths = {
        'A': 8,   # S.No
        'B': 30,  # Name
        'C': 10,  # Rating
        'D': 10,  # Reviews
        'E': 12,  # Price
        'F': 20,  # Category
        'G': 40,  # Address
        'H': 25,  # Hours
        'I': 25,  # Services
        'J': 15   # URL
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 30

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(filename)
    print(f"✅ Excel file created: {filename}")


if __name__ == "__main__":
    extract_coffee_shops()
