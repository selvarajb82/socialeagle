import streamlit as st
import xml.etree.ElementTree as ET
import pandas as pd
from collections import defaultdict
from io import BytesIO
import os

from openpyxl.styles import Font, PatternFill, Border, Side

st.set_page_config(page_title="Payment XML → Excel", layout="wide")
st.title("💳 Payment XML to Excel Converter")

uploaded_file = st.file_uploader(
    "Upload payment XML file (.xml or .txt)",
    type=["xml", "txt"]
)

if uploaded_file:
    try:
        tree = ET.parse(uploaded_file)
        root = tree.getroot()
        ns = {"ns": root.tag.split("}")[0].strip("{")}

        transactions = []
        summary = defaultdict(lambda: {"count": 0, "sum": 0.0})

        # -------- READ GRP HDR --------
        grp_nb = root.find(".//ns:GrpHdr/ns:NbOfTxs", ns)
        grp_sum = root.find(".//ns:GrpHdr/ns:CtrlSum", ns)

        grp_df = pd.DataFrame([{
            "Group Header NbOfTxs": int(grp_nb.text) if grp_nb is not None else "",
            "Group Header CtrlSum": float(grp_sum.text) if grp_sum is not None else ""
        }])

        # -------- PARSE PAYMENTS --------
        for pmt in root.findall(".//ns:PmtInf", ns):
            ppr_id = pmt.findtext(
                "ns:PmtInfId", default="UNKNOWN", namespaces=ns)

            for tx in pmt.findall("ns:CdtTrfTxInf", ns):
                amt_node = tx.find(".//ns:Amt/ns:InstdAmt", ns)
                amount = float(amt_node.text)
                currency = amt_node.attrib.get("Ccy", "")

                creditor = tx.find(".//ns:Cdtr/ns:Nm", ns)
                creditor_name = creditor.text if creditor is not None else ""

                transactions.append({
                    "PPR (PmtInfId)": ppr_id,
                    "Creditor Name": creditor_name,
                    "Amount": amount,
                    "Currency": currency
                })

                summary[ppr_id]["count"] += 1
                summary[ppr_id]["sum"] += amount

        # -------- DATAFRAMES --------
        df_tx = pd.DataFrame(transactions)

        tx_total = df_tx["Amount"].sum()
        df_tx.loc[len(df_tx)] = ["TOTAL", "", tx_total, ""]

        df_summary = pd.DataFrame([
            {
                "PPR (PmtInfId)": k,
                "Number of Transactions": v["count"],
                "Control Sum": round(v["sum"], 2)
            }
            for k, v in summary.items()
        ])

        summary_total_row = {
            "PPR (PmtInfId)": "TOTAL",
            "Number of Transactions": df_summary["Number of Transactions"].sum(),
            "Control Sum": df_summary["Control Sum"].sum()
        }
        df_summary = pd.concat([df_summary, pd.DataFrame([summary_total_row])])

        # -------- STREAMLIT DISPLAY --------
        st.subheader("📄 Transactions")
        st.dataframe(df_tx, use_container_width=True)

        st.subheader("📊 Summary (PPR-wise)")
        st.dataframe(df_summary, use_container_width=True)

        st.subheader("📌 Group Header Summary")
        st.dataframe(grp_df, use_container_width=True)

        # -------- EXCEL EXPORT --------
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_tx.to_excel(writer, sheet_name="Transactions", index=False)
            df_summary.to_excel(writer, sheet_name="Summary",
                                index=False, startrow=0)
            # grp_df.to_excel(writer, sheet_name="Summary",
            #                 index=False, startrow=len(df_summary)+3)

            grp_df.to_excel(writer,  sheet_name="Summary",
                            index=False,
                            startrow=len(df_summary) + 3,
                            startcol=1   # 👈 shifts Group Header table to column B
                            )

            wb = writer.book

            header_fill = PatternFill("solid", fgColor="C6E0B4")
            total_fill = PatternFill("solid", fgColor="FFF2CC")

            header_font = Font(bold=True, size=13)
            total_font = Font(bold=True)

            border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )

            for sheet in wb.worksheets:
                for col in sheet.columns:
                    sheet.column_dimensions[col[0].column_letter].width = 24

                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border

                for row in sheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border

                for row in sheet.iter_rows():
                    if row[0].value == "TOTAL":
                        for cell in row:
                            cell.font = total_font
                            cell.fill = total_fill

        output.seek(0)

        base_name = os.path.splitext(uploaded_file.name)[0]
        excel_name = f"{base_name}_extract.xlsx"

        st.download_button(
            "⬇️ Download Excel File",
            data=output,
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Error processing file: {e}")
